from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from datetime import datetime
import calendar
import os

class ExcelGenerator:
    OUTPUT_PATH = "D:/Project/VSCode/OT_app/data/excel/"
    DEPT_CODE = "K390140R1C"
    DEPT_NAME = "BG6-RD Center-Automatic System Test R&D Div.1-Dept.1-PTB Sec.1"

    @classmethod
    def generate_excel_files(cls, data, date_obj):
        os.makedirs(cls.OUTPUT_PATH, exist_ok=True)
        
        excel_filename = f"{cls.OUTPUT_PATH}{date_obj.strftime('%Y%m%d')}OT.xlsx"
        summary_filename = f"{cls.OUTPUT_PATH}{date_obj.strftime('%Y%m%d')}OTSummary.xlsx"
        
        cls.create_ot_form(excel_filename, data, date_obj)
        cls.create_ot_summary(summary_filename, data, date_obj)
        
        return excel_filename, summary_filename
    
    @classmethod
    def create_ot_form(cls, filename, data, date_info):
        wb = Workbook()
        ws = wb.active
    
        # Column widths setup
        columns = {
            'A': 4.14, 'B': 13.43, 'C': 21.14, 'D': 22.29, 'E': 22.86,
            'F': 10.14, 'G': 15.14, 'H': 13.57, 'I': 15.71, 'J': 1.00,
            'K': 17.14, 'L': 14.43, 'M': 15.43, 'N': 18.57
        }
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 80
        ws.row_dimensions[3].height = 92

        # Title
        ws.merge_cells('A1:N1')
        ws['A1'] = 'Form Lembur (加 班 申 请 单)'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font = Font(name='Arial', size=20)

        # Department info
        ws.merge_cells('A2:C2')
        ws['A2'] = f'Departemen\n(部门代码)：\n{cls.DEPT_CODE}\n{cls.DEPT_NAME}'
        ws['A2'].alignment = Alignment(vertical='center', wrap_text=True)
        ws['A2'].font = Font(name='Arial', size=11)

        # Employee classification
        ws['D2'] = 'Klasifikasi Karyawan\n(人员分类):'
        ws['D2'].alignment = Alignment(vertical='center', horizontal='right', wrap_text=True)
        ws['E2'] = '☐ Band0(1-2職等)\n☑ Band1(3-5職等)'
        ws['E2'].alignment = Alignment(vertical='center', wrap_text=True)

        # Date information
        ch_date = f"{date_info.year} 年 {date_info.month:02d} 月 {date_info.day:02d} 日"
        weekday = calendar.day_name[date_info.weekday()].upper()
        ws.merge_cells('F2:H2')
        ws['F2'] = f'Tanggal Lembur （加班日期）:\n                           {ch_date}\nHari （星期）: {weekday}'
        ws['F2'].alignment = Alignment(vertical='center', wrap_text=True)

        print("Excel data received:", data)
        print("First request:", data[0] if data else "No data")

        # Add overtime type based on is_weekend/is_holiday
        overtime_type = 'Jenis Lembur （加班类别） :\n'
        if data and len(data) > 0:
            first_request = data[0]
            is_weekend = first_request.get('is_weekend', False)
            is_holiday = first_request.get('is_holiday', False)

            if is_holiday:
                overtime_type += ('☐ Saat Hari Kerja (工作日延长加班)\n'
                                '☐ Saat Hari Libur (休息日加班)\n'
                                '☑ Saat Tanggal Merah (法定假日加班)')
            elif is_weekend:
                overtime_type += ('☐ Saat Hari Kerja (工作日延长加班)\n'
                                '☑ Saat Hari Libur (休息日加班)\n'
                                '☐ Saat Tanggal Merah (法定假日加班)')
            else:
                overtime_type += ('☑ Saat Hari Kerja (工作日延长加班)\n'
                                '☐ Saat Hari Libur (休息日加班)\n'
                                '☐ Saat Tanggal Merah (法定假日加班)')
        
        ws.merge_cells('I2:N2')
        ws['I2'] = overtime_type
        ws['I2'].alignment = Alignment(vertical='center', wrap_text=True)

        # Column headers
        headers = [
            ('A3', 'No'),
            ('B3', 'No Karyawan\n(工号)'),
            ('C3', 'Nama\n(姓名)'),
            ('D3', 'Alasan Lembur\n(申请加班事由)'),
            ('E3', 'Jam Lembur (Waktu)\n(预计加班\n起止时间)'),
            ('F3', 'Durasi Lembur\n(预计加班时数)'),
            ('G3', 'Jam istirahat saat lembur (jika perlu, gunakan V)\n(预计休息或用餐打V)'),
            ('H3', 'Tanda Tangan Karyawan\n(员工签名)'),
            ('I3', 'Tanda Tangan Supervisor\n(课级主管签名)'),
            ('K3', 'Konfirmasi Jam Lembur (Waktu)\n(实际加班\n起止时间)'),
            ('L3', 'Konfirmasi Durasi Lembur\n(实际加班时数)'),
            ('M3', 'Jam Istirahat (Jika ada, gunakan V)\n(实际休息或用餐打V)'),
            ('N3', 'Tanda Tangan Karyawan\n(员工签名)')
        ]

        # Apply header formatting
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        header_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_font = Font(name='Arial', size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font
            ws[cell].border = thin_border

        # Add empty rows with formatting
        for row in range(4, 34):
            for col in 'ABCDEFGHIKLMN':
                cell = f'{col}{row}'
                ws[cell] = ''
                ws[cell].border = thin_border
                ws[cell].font = Font(name='Arial', size=11)
                ws[cell].alignment = Alignment(vertical='center')
            ws[f'A{row}'] = row - 3  # Add row numbers

        # Fill in data
        current_row = 4
        for item in data:
            ws[f'B{current_row}'] = item['employee_id']
            ws[f'C{current_row}'] = item['employee_name']
            ws[f'D{current_row}'] = item['reason']
            ws[f'E{current_row}'] = f"{item['time_start']} - {item['time_end']}"
            ws[f'F{current_row}'] = f"{item['total_hours']} hour(s)"
            ws[f'G{current_row}'] = 'V' if item['has_break'] else '-'
            
            # Add actual time, duration and break status
            ws[f'K{current_row}'] = f"{item['time_start']} - {item['time_end']}"
            ws[f'L{current_row}'] = f"{item['total_hours']} hour(s)"
            ws[f'M{current_row}'] = 'V' if item['has_break'] else '-'
            
            current_row += 1

        # Footer notes
        ws.merge_cells('A34:N34')
        ws['A34'] = 'Keterangan：'
        ws['A34'].font = Font(name='Arial', size=12)

        notes = [
            ('A35:N35', '1、Informasi diatas harus dilaporkan dengan benar, pelanggaran akan dikenakan sesuai dengan hukuman yang ada dari managemen perusahaan\n(以上资料请据实申报，违者按奖惩管理办法处理)。'),
            ('A36:N36', '2、Karyawan harus menandatangani aplikasi lembur. Jika tidak ada tanda tangan maka akan di anggap tidak sah (实际加班时数，以员工签名确认为准)。'),
            ('A37:N37', '3、Tidak ada aplikasi lembur tidak akan di hitung untuk upah lembur (无加班申请单不计发加班费)。')
        ]

        for cells, text in notes:
            ws.merge_cells(cells)
            ws[cells.split(':')[0]] = text
            ws[cells.split(':')[0]].font = Font(name='Arial', size=12)
            ws[cells.split(':')[0]].alignment = Alignment(vertical='center', wrap_text=True)

        # Form number
        ws['M38'] = 'Form No.:PTB-TB004-001 Rev.01'
        ws['M38'].font = Font(name='Arial', size=10)

        # Save workbook
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)

    @classmethod
    def create_ot_summary(cls, filename, data, date_info):
        wb = Workbook()
        ws = wb.active
        
        # Set column widths
        columns = {
            'A': 11.5,  # Work ID
            'B': 15.5,  # Name
            'C': 21.0,  # Project
            'D': 9.75,  # Start Time
            'E': 9.75,  # End Time
            'F': 9.75,  # Break
            'G': 6.0,   # Hours
            'H': 25.0   # Reason
        }
        
        for col, width in columns.items():
            ws.column_dimensions[col].width = width

        # Headers with formatting
        headers = [
            ('A1', 'Work ID'),
            ('B1', 'Name'),
            ('C1', 'Project'),
            ('D1', 'Start Time'),
            ('E1', 'End Time'),
            ('F1', 'Break'),
            ('G1', 'Hours'),
            ('H1', 'Reason')
        ]

        header_style = Alignment(horizontal='center', vertical='center')
        header_font = Font(name='Arial', size=11)

        for cell, text in headers:
            ws[cell] = text
            ws[cell].alignment = header_style
            ws[cell].font = header_font

        # Fill data
        current_row = 2
        for item in data:
            ws[f'A{current_row}'] = item['employee_id']
            ws[f'B{current_row}'] = item['employee_name']
            ws[f'C{current_row}'] = item['project']
            ws[f'D{current_row}'] = item['time_start']
            ws[f'E{current_row}'] = item['time_end']
            ws[f'F{current_row}'] = 'Y' if item['has_break'] else 'N'
            ws[f'G{current_row}'] = item['total_hours']
            ws[f'H{current_row}'] = item['reason']

            # Formatting
            for col in 'ABCDEFGH':
                cell = f'{col}{current_row}'
                ws[cell].font = Font(name='Arial', size=11)
                ws[cell].alignment = Alignment(vertical='center')
                if col in 'ADEFG':
                    ws[cell].alignment = Alignment(horizontal='center', vertical='center')

            current_row += 1

        # Save
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)